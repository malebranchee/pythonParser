from bs4 import BeautifulSoup 
import requests 
from openpyxl.reader.excel import load_workbook



#Создание книги Excel



def parse():

    url = 'https://www.chitai-gorod.ru/' #Запрос на сайт
    page = requests.get(url) 
    print(page.status_code)
    soup = BeautifulSoup(page.text, "html.parser") 


    allAuthors = soup.findAll('div', class_='product-title__author') 
    allNames = soup.findAll('div', class_='product-title__head' )
    allPrices = soup.findAll('div', class_='product-price__value')
    sortExcel(allAuthors, allNames, allPrices)

def sortExcel(authors, names, prices): #Функция заполнения ячеек Excel

    doc = load_workbook(filename='goo.xlsx')
    ws = doc.active #Создание книги Excel и рабочего листа

    i = 1
    j = 1
    for auth in authors:
        n_auth = " ".join(auth.text.split())
        if n_auth is not None:
            ws.cell(row=j, column =i).value = n_auth
            j += 1
    i += 1
    j = 1
    for name in names:
        n_name = " ".join(name.text.split())
        if n_name is not None:
            ws.cell(row=j, column = i).value = n_name
            j += 1
    i += 1
    j = 1
    for price in prices:
        n_price = " ".join(price.text.split())
        if n_price is not None:
            ws.cell(row=j, column = i).value = n_price
            j += 1

    doc.save('goo.xlsx')






